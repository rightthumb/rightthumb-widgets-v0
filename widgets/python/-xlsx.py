import _rightThumb._construct as __;appDBA=__.clearFocus(__name__,__file__);__.appReg=appDBA;import _rightThumb._base3 as _; # type: ignore
def focus(parentApp='', childApp='', reg=True): global appDBA; f = __.appName(appDBA, parentApp, childApp); return f if reg else f
fieldSet=_.l.vars(focus(),__name__,__file__,appDBA);_.load();_v=__.imp('_rightThumb._vars')

def sw():
    pass
    _.switches.register( 'Files', '-f,-fi,-file,-files','file.txt  ||  folderAlias==FileInThatFolder.txt', isData='glob', description='Files', isRequired=False )
_._default_settings_()

_.appInfo[focus()] = {
    'file': 'thisApp.py',
    'description': 'Changes the world',
    'categories': [
                        'DEFAULT',
                ],
    'examples': [
                        _.hp('p thisApp -file file.txt'),
                        _.linePrint(label='simple',p=0),
                        '',
    ],
    'columns': [
    ],
    'aliases': [],
    'relatedapps': [],
    'prerequisite': [],
    'notes': [],
}

_.appInfo[focus()] = _.appInfoContinuity(__.thisApp( __file__ ),_.appInfo[focus()])
_.appData[focus()] = _.appDataContinuity()
def appRegDics(): return { 'appInfo': _.appInfo[focus()], 'appData': _.appData[focus()] }

def triggers():
    _._default_triggers_()
    _.switches.trigger( 'Files',   _.isFileAdvanced, vs=False )     # Advanced File Registration    (Fn Alias Resolves To: def myFileLocations)
    _.switches.trigger( 'DB', _.aliasesFi )
    _.switches.trigger( 'Folder', _.myFolderLocations )
    _.switches.trigger( 'Folders', _.myFolderLocations )
    __.SwitchesModifier.Trigger['Folders'] = _.myFolder
    _.switches.trigger( 'OutputFolder', _.aliasesFo )
def _local_(do): exec(do)
_.l.conf('clean-pipe',True); _.l.sw.register( triggers, sw )
########################################################################################
#n)--> start

def spreadsheet():
    from openpyxl import load_workbook
    from datetime import datetime, time
    FIELDS = ["date", "time", "status", "name", "phone", "message", "kind"]

    def normalize(v):
        if isinstance(v, datetime):
            return v.isoformat()
        if isinstance(v, time):
            return v.isoformat()
        return v

    def is_valid_row(row):
        if not row:
            return False
        if isinstance(row[0], str) and row[0].lower().startswith("exported on"):
            return False
        if all(v is None for v in row):
            return False
        return (
            isinstance(row[0], datetime) and
            isinstance(row[1], time) and
            row[2] in ("Sent", "Received")
        )
    wb = load_workbook("sms.xlsx")
    ws = wb.active
    records = []
    for row in ws.iter_rows(values_only=True):
        if not is_valid_row(row):
            continue
        rec = {}
        for i, key in enumerate(FIELDS):
            val = row[i] if i < len(row) else None
            rec[key] = normalize(val)
        records.append(rec)
    import re

    def extract_numbers_by_length(text, lengths=None):
        if text in (None, "", "nan"):
            return []
        text = str(text)
        nums = re.findall(r'\d+', text)
        if lengths:
            nums = [n for n in nums if len(n) in lengths]
        return nums
    lengths = {6,7, 16,17, 20}
    jobs = []

    justJobs = True
    codes = {}
    instas = {}
    for rec in records:
        msg = rec.get("message", "")
        rec["numbers"] = extract_numbers_by_length(msg, lengths)
        
        try:
            msg.splitlines()
        except: continue

        if rec["numbers"]:
            jobType = ''
            insta = ''

            if not justJobs:
                _.pr(line=1 )
            for line in msg.splitlines():
                if ' > ' in line:
                    jobType = line.strip().split(' > ')[0].strip()
                    insta = line.strip().split(' > ')[1].strip()
                    break
            for j in extract_numbers_by_length(msg, {6,7}):
                print(j)
                if jobType:
                    codes[j] = jobType
                    instas[j] = insta
            if not justJobs:
                print(rec["message"])
            jobs.append(rec)
    # asdf
    for file in _.switches.values('Files'):
        contents = _.getText(file, raw=True)
        rec = {}
        rec["numbers"] = extract_numbers_by_length(msg, {6,7})
        jobType = ''
        insta = ''
        if rec["numbers"]:
            for line in contents.splitlines():
                if ' > ' in line:
                    jobType = line.strip().split(' > ')[0].strip()
                    insta = line.strip().split(' > ')[1].strip()
                    break
            for j in extract_numbers_by_length(contents, {6,7}):
                print(j)
                if jobType:
                    codes[j] = jobType
                    instas[j] = insta
    
    _.saveTable2(codes, 'live.json')
    _.saveTable2(instas, 'types.json')
    if not justJobs:
        print()
        print()
        print()
        print("TOTAL:", len(records))
        print("JOBS:", len(jobs))


def getFieldIndexes(headerRow, fieldNames=[]):
    fieldNames = [
        'Tech #',
        'Job #',
        'Date',
        'Address',
        'Job Code',
        'Qty',
        'Rate',
        'Total'
    ]


    indexes = {}
    for field in fieldNames:
        try:
            idx = headerRow.index(field)
            indexes[field] = idx
        except ValueError:
            indexes[field] = None
    return indexes
def buildFieldRanges(headerLine, fieldNames=[]):
    fieldNames = [
        'Tech #',
        'Job #',
        'Date',
        'Address',
        'Job Code',
        'Qty',
        'Rate',
        'Total'
    ]

    ranges = {}
    
    for field in fieldNames:
        start = headerLine.find(field)
        if start == -1:
            ranges[field] = None
            continue
        
        end = start + len(field)
        
        # expand right until next non-space gap (next column start)
        while end < len(headerLine) and headerLine[end] == ' ':
            end += 1
        
        ranges[field] = (start, end)
    
    return ranges



def indexPaycheckByJobs():
    file = _.getText('all_paychecks.all', raw=True).splitlines()
    jobs = []
    for line in file:
        if 'Job Code' in line and 'Tech #' in line:
            
            print(line.index('Job Code'))

def action():
    # indexPaycheckByJobs()
    spreadsheet()

########################################################################################
if __name__ == '__main__':
    action(); _.isExit(__file__)